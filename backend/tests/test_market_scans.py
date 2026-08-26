import uuid
from datetime import UTC, date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from mkvip.api.dependencies import get_market_scan_repository
from mkvip.api.routes.market_scans import get_market_scan_executor
from mkvip.db.base import Base
from mkvip.main import app
from mkvip.models.user import UserOrm
from mkvip.providers.base import (
    ProviderCompanySearchResult,
    ProviderDataError,
    ProviderPricePoint,
)
from mkvip.providers.market_universe import MarketSecurity, MKVIPIndexUniverseProvider
from mkvip.repositories.market_scan import SqlAlchemyMarketScanRepository
from mkvip.schemas.index import IndexCompositionRead, IndexConstituentRead, IndexSummaryRead
from mkvip.schemas.market_scan import MarketScanCriteria, MarketScanRead, MarketScanResultRead
from mkvip.services.market_scan_export import build_market_scan_workbook
from mkvip.services.market_scans import MarketScanService, criteria_from_question


class RecordingRepository:
    def __init__(self) -> None:
        self.total = 0
        self.status = "queued"
        self.results = []
        self.processed = 0
        self.matched = 0
        self.failed = 0
        self.insufficient = 0
        self.error = None

    async def mark_running(self, scan_id, total) -> bool:
        self.status = "running"
        self.total = total
        return True

    async def record_batch(
        self,
        scan_id,
        results,
        *,
        processed,
        matched,
        failed,
        insufficient,
    ) -> bool:
        self.results.extend(results)
        self.processed = processed
        self.matched = matched
        self.failed = failed
        self.insufficient = insufficient
        return True

    async def mark_completed(self, scan_id) -> None:
        self.status = "completed"

    async def mark_failed(self, scan_id, message) -> None:
        self.status = "failed"
        self.error = message


class ApiRepository:
    def __init__(self) -> None:
        self.scan = None

    async def list(self):
        return [] if self.scan is None else [self.scan]

    async def create(self, criteria, request_text):
        now = datetime.now(UTC)
        self.scan = MarketScanRead(
            id=uuid.uuid4(),
            status="queued",
            criteria=criteria,
            request_text=request_text,
            universe_source="Nasdaq public screener",
            price_source="Yahoo Finance",
            total_securities=0,
            processed_securities=0,
            matched_securities=0,
            failed_securities=0,
            insufficient_history_securities=0,
            progress_pct=0,
            error_message=None,
            created_at=now,
            started_at=None,
            completed_at=None,
            results=[],
        )
        return self.scan

    async def get(self, scan_id):
        return self.scan if self.scan and self.scan.id == scan_id else None

    async def cancel(self, scan_id):
        if self.scan is None or self.scan.id != scan_id:
            return None
        self.scan = self.scan.model_copy(
            update={"status": "cancelled", "completed_at": datetime.now(UTC)}
        )
        return self.scan


class UniverseProvider:
    name = "Test universe"

    async def list_us_equities(self, exchanges):
        return [
            MarketSecurity("DROP", "Drop Corporation", "NASDAQ", "US", "USD", 2e9),
            MarketSecurity("FLAT", "Flat Corporation", "NYSE", "US", "USD", 1e9),
            MarketSecurity("FAIL", "Failure Corporation", "AMEX", "US", "USD", 1e8),
            MarketSecurity("SPACW", "Example Warrants", "NASDAQ", "US", "USD", 1e8),
        ]


class PriceProvider:
    name = "Public prices"

    async def get_price_history(self, ticker):
        if ticker == "FAIL":
            raise ProviderDataError("temporary")
        end = 10 if ticker == "DROP" else 70
        return [
            ProviderPricePoint("2021-08-26", 100, 100),
            ProviderPricePoint("2026-08-26", end, end),
        ]


class BatchPriceProvider(PriceProvider):
    def __init__(self) -> None:
        self.batch_calls = []

    async def get_price_histories(self, tickers, years):
        self.batch_calls.append((tickers, years))
        return {
            ticker: await super().get_price_history(ticker)
            for ticker in tickers
            if ticker != "FAIL"
        }

    async def get_price_history(self, ticker):
        raise AssertionError("Le chemin individuel ne doit pas être utilisé.")


class IndexUniverse:
    name = "MK-VIP indices"

    def __init__(self) -> None:
        self.codes = []

    async def list_index_equities(self, index_code):
        self.codes.append(index_code)
        return [MarketSecurity("DROP", "Drop Corporation", "XPAR", "France", "EUR", None)]


@pytest.mark.asyncio
async def test_scan_filters_and_calculates_five_year_decline() -> None:
    repository = RecordingRepository()
    service = MarketScanService(
        repository,
        UniverseProvider(),
        PriceProvider(),
        concurrency=2,
        retry_delay_seconds=0,
    )

    await service.run(uuid.uuid4(), MarketScanCriteria())

    assert repository.status == "completed"
    assert repository.total == 3
    assert repository.processed == 3
    assert repository.matched == 1
    assert repository.failed == 1
    assert repository.insufficient == 0
    assert repository.results[0].ticker == "DROP"
    assert repository.results[0].performance_pct == -90


@pytest.mark.asyncio
async def test_full_market_scan_uses_batched_price_downloads() -> None:
    repository = RecordingRepository()
    provider = BatchPriceProvider()
    service = MarketScanService(
        repository,
        UniverseProvider(),
        provider,
        concurrency=2,
        retry_delay_seconds=0,
    )

    await service.run(uuid.uuid4(), MarketScanCriteria())

    assert provider.batch_calls == [(["DROP", "FLAT", "FAIL"], 5)]
    assert repository.status == "completed"
    assert repository.processed == 3
    assert repository.matched == 1
    assert repository.insufficient == 1


@pytest.mark.asyncio
async def test_scan_can_use_an_mkvip_index_instead_of_the_full_us_market() -> None:
    repository = RecordingRepository()
    index_universe = IndexUniverse()
    service = MarketScanService(
        repository,
        UniverseProvider(),
        PriceProvider(),
        index_universe_provider=index_universe,
        retry_delay_seconds=0,
    )

    await service.run(
        uuid.uuid4(),
        MarketScanCriteria(market="INDEX", index_code="cac-40"),
    )

    assert index_universe.codes == ["CAC40"]
    assert repository.status == "completed"
    assert repository.total == 1
    assert repository.results[0].ticker == "DROP"


@pytest.mark.asyncio
async def test_index_universe_builds_yahoo_symbols_for_local_exchanges() -> None:
    class Catalog:
        async def get_composition(self, code):
            return IndexCompositionRead(
                code=code,
                name="Test Europe",
                market="Europe",
                provider="Test",
                region="Europe",
                country="France",
                source_url="https://example.com",
                constituents=[
                    IndexConstituentRead(
                        name="SAP",
                        ticker="SAP",
                        mic="XETR",
                        trading_location="Xetra",
                        country="Allemagne",
                        currency="EUR",
                    ),
                    IndexConstituentRead(
                        name="L'Oréal",
                        isin="FR0000120321",
                        mic="XPAR",
                        trading_location="Euronext Paris",
                        country="France",
                        currency="EUR",
                    ),
                ],
            )

    class Discovery:
        async def search_company(self, query):
            return [ProviderCompanySearchResult("OR.PA", "L'Oréal", "Paris")]

    securities = await MKVIPIndexUniverseProvider(Catalog(), Discovery()).list_index_equities(
        "TEST"
    )

    assert [security.ticker for security in securities] == ["SAP.DE", "OR.PA"]


def test_agent_question_is_converted_to_verified_criteria() -> None:
    criteria = criteria_from_question(
        "Trouve sur le NASDAQ les actions en baisse d’au moins 85 % sur 3 ans "
        "avec une capitalisation de 1 milliard"
    )

    assert criteria.exchanges == ["NASDAQ"]
    assert criteria.minimum_decline_pct == 85
    assert criteria.years == 3
    assert criteria.minimum_market_cap == 1_000_000_000


def test_agent_question_recognizes_any_mkvip_index_name_or_code() -> None:
    indices = [
        IndexSummaryRead(
            code="SP500",
            name="S&P 500",
            market="XNYS",
            provider="iShares",
            region="Amérique",
            country="États-Unis",
        ),
        IndexSummaryRead(
            code="EUROPEBANKS",
            name="STOXX Europe 600 Banks",
            market="Europe",
            provider="SPDR",
            region="Europe",
            country="Europe",
            kind="sector",
            sector="Financials",
        ),
    ]

    by_name = criteria_from_question(
        "Quelles actions du S&P 500 ont baissé de 70 % sur 5 ans ?", indices
    )
    by_code = criteria_from_question(
        "Analyse EUROPEBANKS sur 3 ans avec une baisse de 60 %", indices
    )

    assert (by_name.market, by_name.index_code) == ("INDEX", "SP500")
    assert (by_code.market, by_code.index_code) == ("INDEX", "EUROPEBANKS")
    assert by_code.years == 3


def test_completed_scan_can_be_exported_as_a_readable_workbook() -> None:
    now = datetime(2026, 8, 26, tzinfo=UTC)
    scan = MarketScanRead(
        id=uuid.uuid4(),
        status="completed",
        criteria=MarketScanCriteria(),
        request_text="Actions US en baisse de 80 % sur 5 ans",
        universe_source="Nasdaq public screener",
        price_source="Yahoo Finance",
        total_securities=1,
        processed_securities=1,
        matched_securities=1,
        failed_securities=0,
        insufficient_history_securities=0,
        progress_pct=100,
        error_message=None,
        created_at=now,
        started_at=now,
        completed_at=now,
        results=[
            {
                "id": uuid.uuid4(),
                "ticker": "DROP",
                "name": "Drop Corporation",
                "exchange": "NASDAQ",
                "country": "États-Unis",
                "currency": "USD",
                "market_cap": 2_000_000_000,
                "start_date": date(2021, 8, 26),
                "end_date": date(2026, 8, 26),
                "start_price": 100,
                "end_price": 10,
                "performance_pct": -90,
                "price_source": "Yahoo Finance",
            }
        ],
    )

    workbook = load_workbook(BytesIO(build_market_scan_workbook(scan)))

    assert workbook.sheetnames == ["Synthèse", "Résultats"]
    assert workbook["Résultats"]["C2"].value == "DROP"
    assert workbook["Résultats"]["K2"].value == -0.9


@pytest.mark.asyncio
async def test_scan_progress_and_results_are_persisted_per_owner() -> None:
    engine = create_async_engine("sqlite+aiosqlite://")
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        owner = UserOrm(email="scanner@example.com", password_hash="unused")
        session.add(owner)
        await session.commit()
        repository = SqlAlchemyMarketScanRepository(session, owner.id)
        scan = await repository.create(MarketScanCriteria(), "baisse de 80 % sur 5 ans")
        await repository.mark_running(scan.id, 10)
        await repository.record_batch(
            scan.id,
            [
                MarketScanResultRead(
                    id=uuid.uuid4(),
                    ticker="DROP",
                    name="Drop Corporation",
                    exchange="NASDAQ",
                    country="US",
                    currency="USD",
                    market_cap=1_000_000,
                    start_date=date(2021, 8, 26),
                    end_date=date(2026, 8, 26),
                    start_price=100,
                    end_price=10,
                    performance_pct=-90,
                    price_source="Yahoo Finance",
                )
            ],
            processed=5,
            matched=1,
            failed=0,
            insufficient=1,
        )
        persisted = await repository.get(scan.id)
        assert persisted is not None
        assert persisted.progress_pct == 50
        assert persisted.results[0].ticker == "DROP"
    await engine.dispose()


@pytest.mark.asyncio
async def test_cancelled_scan_cannot_be_completed_by_the_background_worker() -> None:
    engine = create_async_engine("sqlite+aiosqlite://")
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as worker_session, factory() as control_session:
        owner = UserOrm(email="cancel@example.com", password_hash="unused")
        worker_session.add(owner)
        await worker_session.commit()
        worker = SqlAlchemyMarketScanRepository(worker_session, owner.id)
        control = SqlAlchemyMarketScanRepository(control_session, owner.id)
        scan = await worker.create(MarketScanCriteria(), "scan à arrêter")
        assert await worker.mark_running(scan.id, 100)

        cancelled = await control.cancel(scan.id)
        await worker.mark_completed(scan.id)
        persisted = await worker.get(scan.id)

        assert cancelled is not None
        assert cancelled.status == "cancelled"
        assert persisted is not None
        assert persisted.status == "cancelled"
        assert persisted.progress_pct == 0
    await engine.dispose()


def test_agent_endpoint_creates_a_background_scan(client) -> None:
    repository = ApiRepository()
    executions = []

    async def execute(scan_id, owner_id) -> None:
        executions.append((scan_id, owner_id))

    app.dependency_overrides[get_market_scan_repository] = lambda: repository
    app.dependency_overrides[get_market_scan_executor] = lambda: execute

    response = client.post(
        "/api/v1/market-scans/from-question",
        json={"question": "Actions du NASDAQ en baisse de 85 % sur 3 ans"},
    )

    assert response.status_code == 202
    payload = response.json()
    assert payload["status"] == "queued"
    assert payload["criteria"]["exchanges"] == ["NASDAQ"]
    assert payload["criteria"]["minimum_decline_pct"] == 85
    assert payload["criteria"]["years"] == 3
    assert executions and executions[0][0] == repository.scan.id


def test_running_scan_can_be_cancelled(client) -> None:
    repository = ApiRepository()

    async def execute(scan_id, owner_id) -> None:
        return None

    app.dependency_overrides[get_market_scan_repository] = lambda: repository
    app.dependency_overrides[get_market_scan_executor] = lambda: execute
    created = client.post(
        "/api/v1/market-scans",
        json={"criteria": {"years": 5, "minimum_decline_pct": 80}},
    )

    response = client.post(
        f"/api/v1/market-scans/{created.json()['id']}/cancel",
    )

    assert response.status_code == 200
    assert response.json()["status"] == "cancelled"
