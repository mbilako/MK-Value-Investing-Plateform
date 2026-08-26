from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mkvip.core.national_markets import get_national_market
from mkvip.schemas.market_scan import MarketScanRead


def build_market_scan_workbook(scan: MarketScanRead) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Synthèse"
    is_index_scan = scan.criteria.market == "INDEX"
    national_market = get_national_market(scan.criteria.country_code)
    universe = (
        scan.criteria.index_code
        if is_index_scan
        else f"Marché national — {national_market.name}"
        if national_market is not None
        else "Marché américain"
    )
    places = (
        "Indice MK-VIP"
        if is_index_scan
        else ", ".join(national_market.yahoo_exchanges)
        if national_market is not None
        else ", ".join(scan.criteria.exchanges)
    )
    rows = [
        ("MK-VIP — Scan de marché", None),
        ("Statut", scan.status),
        ("Univers", universe),
        ("Places", places),
        ("Période", f"{scan.criteria.years} ans"),
        ("Baisse minimale", scan.criteria.minimum_decline_pct / 100),
        ("Capitalisation minimale", scan.criteria.minimum_market_cap),
        ("Titres examinés", scan.processed_securities),
        ("Résultats", scan.matched_securities),
        ("Historiques insuffisants", scan.insufficient_history_securities),
        ("Échecs de source", scan.failed_securities),
        ("Source de l’univers", scan.universe_source),
        ("Source des cours", scan.price_source),
        ("Créé le", _excel_datetime(scan.created_at)),
        ("Terminé le", _excel_datetime(scan.completed_at)),
    ]
    for row in rows:
        summary.append(row)
    summary["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="123D35")
    summary.merge_cells("A1:B1")
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 34
    summary["B6"].number_format = "0.0%"
    summary.freeze_panes = "A2"

    results = workbook.create_sheet("Résultats")
    headers = [
        "Pays",
        "Place",
        "Ticker",
        "Entreprise",
        "Capitalisation",
        "Devise",
        "Date de départ",
        "Cours de départ",
        "Date de fin",
        "Cours de fin",
        "Performance",
        "Source",
    ]
    results.append(headers)
    for item in scan.results:
        results.append(
            [
                item.country,
                item.exchange,
                item.ticker,
                item.name,
                item.market_cap,
                item.currency,
                item.start_date,
                item.start_price,
                item.end_date,
                item.end_price,
                item.performance_pct / 100,
                item.price_source,
            ]
        )
    for cell in results[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123D35")
        cell.alignment = Alignment(horizontal="center")
    results.freeze_panes = "A2"
    results.auto_filter.ref = results.dimensions
    widths = [18, 12, 14, 42, 18, 10, 16, 16, 16, 16, 14, 18]
    for index, width in enumerate(widths, start=1):
        results.column_dimensions[get_column_letter(index)].width = width
    for row in results.iter_rows(min_row=2):
        row[4].number_format = '#,##0'
        row[7].number_format = '0.0000'
        row[9].number_format = '0.0000'
        row[10].number_format = '0.00%'

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_datetime(value):
    return value.replace(tzinfo=None) if value is not None and value.tzinfo else value
